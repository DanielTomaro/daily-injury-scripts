import os
import requests
import pandas as pd
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

# ESPN API Base URLs for different sports
SPORTS_API_URLS = {
    "NBA": "https://sports.core.api.espn.com/v2/sports/basketball/leagues/nba/teams/{}/injuries",
    "NFL": "https://sports.core.api.espn.com/v2/sports/football/leagues/nfl/teams/{}/injuries",
    "MLB": "https://sports.core.api.espn.com/v2/sports/baseball/leagues/mlb/teams/{}/injuries",
    "NHL": "https://sports.core.api.espn.com/v2/sports/hockey/leagues/nhl/teams/{}/injuries"
}

# Team IDs for each sport
SPORT_TEAM_IDS = {
    "NBA": {
        "Atlanta Hawks": 1, "Boston Celtics": 2, "Brooklyn Nets": 17, "Chicago Bulls": 4,
        "Los Angeles Lakers": 13, "Miami Heat": 14, "New York Knicks": 18, "Golden State Warriors": 9
    },
    "NFL": {
        "Buffalo Bills": 2, "Dallas Cowboys": 6, "San Francisco 49ers": 25, "Kansas City Chiefs": 12
    },
    "MLB": {
        "New York Yankees": 10, "Los Angeles Dodgers": 19, "Boston Red Sox": 2, "Houston Astros": 18
    },
    "NHL": {
        "Toronto Maple Leafs": 21, "Montreal Canadiens": 8, "Boston Bruins": 1, "Chicago Blackhawks": 4
    }
}

# Create output directories
output_dir = "combined_reports"
csv_dir = "league_data"
os.makedirs(output_dir, exist_ok=True)
for league in SPORTS_API_URLS.keys():
    league_dir = os.path.join(csv_dir, f"{league.lower()}_injuries", "latest")
    os.makedirs(league_dir, exist_ok=True)

# Output Excel file with current date
today_date = datetime.today().strftime("%Y-%m-%d")
output_file = os.path.join(output_dir, f"combined_injury_report_{today_date}.xlsx")

# Function to fetch injury details
def get_injury_details(ref_url):
    try:
        response = requests.get(ref_url)
        if response.status_code == 200:
            return response.json()
        else:
            return None
    except Exception as e:
        print(f"❌ Error fetching injury details: {e}")
        return None

# Dictionary to store league data
league_data = {}

# Fetch and process data for each sport
for sport, api_url in SPORTS_API_URLS.items():
    print(f"📥 Fetching {sport} injuries...")
    
    sport_data = []
    
    for team, team_id in SPORT_TEAM_IDS[sport].items():
        url = api_url.format(team_id)
        response = requests.get(url)
        
        if response.status_code == 200:
            data = response.json()
            
            if "items" in data:
                for injury_item in data["items"]:
                    injury_url = injury_item.get("$ref")
                    if injury_url:
                        injury_details = get_injury_details(injury_url)
                        
                        if injury_details:
                            # Extract details
                            player_name = injury_details.get("athlete", {}).get("displayName", "Unknown")
                            position_info = injury_details.get("athlete", {}).get("position", {})
                            position = position_info.get("abbreviation", "N/A")
                            
                            injury_info = injury_details.get("injury", {})
                            injury_type = injury_info.get("type", "N/A")
                            location = injury_info.get("location", "N/A")
                            detail = injury_info.get("detail", "N/A")
                            side = injury_info.get("side", "N/A")
                            
                            # Format dates
                            return_date = injury_info.get("returnDate", "Unknown")
                            reported_date = injury_details.get("date", datetime.today().strftime("%Y-%m-%d"))
                            
                            # Injury status
                            status = injury_info.get("fantasyStatus", {}).get("description", "Unknown")
                            
                            # Comment Format
                            comment = f"{injury_type} ({location}) - {detail} ({side})"
                            
                            # Add to data list
                            sport_data.append({
                                "Team": team,
                                "Player Name": player_name,
                                "Position": position,
                                "Injury Type": injury_type,
                                "Status": status,
                                "Return Date": return_date,
                                "Reported Date": reported_date,
                                "Short Comment": comment
                            })
                            
                        time.sleep(0.5)  # Rate limit
    
    # Save to DataFrame
    if sport_data:
        df = pd.DataFrame(sport_data)
        league_data[sport] = df
        
        # Save to CSV
        csv_file = os.path.join(csv_dir, f"{sport.lower()}_injuries", "latest", f"{sport.lower()}_injuries_latest.csv")
        df.to_csv(csv_file, index=False)
        print(f"✅ Saved {sport} data with {len(df)} injuries")
    else:
        print(f"⚠️ No injury data found for {sport}")

# Create an Excel workbook
wb = Workbook()
# Remove default sheet
default_sheet = wb.active
wb.remove(default_sheet)

# Dictionary to track available leagues for summary
available_leagues = {}
total_injuries = 0

# Create a summary sheet
summary_sheet = wb.create_sheet(title="Summary")
summary_sheet.append(["Combined Injury Report Summary"])
summary_sheet["A1"].font = Font(bold=True, size=16)
summary_sheet.append([])
summary_sheet.append(["League", "Number of Teams with Injuries", "Total Injuries"])
summary_row = 4  # Start after the header

# Process each league
for league, df in league_data.items():
    if not df.empty:
        # Create a new sheet for the league
        ws = wb.create_sheet(title=league)
        
        # Add league name as a bold header
        ws.append([f"{league} Injury Report"])
        ws["A1"].font = Font(bold=True, size=16)
        ws.append([])  # Blank row
        
        # Get unique teams
        teams = sorted(df["Team"].unique())
        teams_with_injuries = len(teams)
        league_injuries = len(df)
        
        # Update total count
        total_injuries += league_injuries
        
        # Add to summary data
        available_leagues[league] = {
            "teams": teams_with_injuries,
            "injuries": league_injuries
        }
        
        row_num = 3  # Start from row 3 after league header
        for team in teams:
            team_data = df[df["Team"] == team]
            if team_data.empty:
                continue
                
            # Add team name as a title
            ws.append([team])
            ws.cell(row=row_num, column=1).font = Font(bold=True, size=14)
            row_num += 1
            
            # Add column headers
            headers = ["Player Name", "Position", "Injury Type", "Status", "Return Date", "Reported Date", "Comment"]
            
            ws.append(headers)
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=row_num, column=col_num).font = Font(bold=True)
            row_num += 1  # Move to the next row
            
            # Add team data
            for _, row in team_data.iterrows():
                row_values = [
                    row["Player Name"],
                    row["Position"],
                    row["Injury Type"],
                    row["Status"],
                    row["Return Date"],
                    row["Reported Date"],
                    row["Short Comment"]
                ]
                
                ws.append(row_values)
                row_num += 1
            
            # Add space before next team
            ws.append([])
            row_num += 1
            
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap width at 50 to avoid extremely wide columns
            ws.column_dimensions[col_letter].width = adjusted_width
            
        print(f"✅ Added {league} with {league_injuries} injuries across {teams_with_injuries} teams")

# Fill in the summary sheet with the data we collected
for league, data in sorted(available_leagues.items()):
    summary_sheet.append([league, data["teams"], data["injuries"]])
    summary_row += 1

# Add total row
summary_sheet.append(["TOTAL", sum(data["teams"] for data in available_leagues.values()), total_injuries])
summary_sheet.cell(row=summary_row, column=1).font = Font(bold=True)
summary_sheet.cell(row=summary_row, column=2).font = Font(bold=True)
summary_sheet.cell(row=summary_row, column=3).font = Font(bold=True)

# Auto-adjust summary sheet column widths
for col in summary_sheet.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    summary_sheet.column_dimensions[col_letter].width = max_length + 2

# Save the Excel file
wb.save(output_file)
print(f"🚀 Combined injury report saved as {output_file}")