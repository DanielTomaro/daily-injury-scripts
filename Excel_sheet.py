import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

# Directory for injury reports - using 'latest' folders within each league's directory
league_files = {
    "MLB": os.path.join("mlb_injuries", "latest", "mlb_injuries_latest.csv"),
    "NHL": os.path.join("nhl_injuries", "latest", "nhl_injuries_latest.csv"),
    "NFL": os.path.join("nfl_injuries", "latest", "nfl_injuries_latest.csv"),
    "NBA": os.path.join("nba_injuries", "latest", "nba_injuries_latest.csv"),
    "AFL": os.path.join("afl_injuries", "latest", "afl_injuries_latest.csv"),
    "NRL": os.path.join("nrl_injuries", "latest", "nrl_injuries_latest.csv")
}

# Create output directory if it doesn't exist
output_dir = "combined_reports"
os.makedirs(output_dir, exist_ok=True)

# Output Excel file with current date
from datetime import datetime
today_date = datetime.today().strftime("%Y-%m-%d")
output_file = os.path.join(output_dir, f"combined_injury_report_{today_date}.xlsx")

# Create an Excel workbook
wb = Workbook()
# Remove default sheet
default_sheet = wb.active
wb.remove(default_sheet)

# Dictionary to track available leagues for summary
available_leagues = {}
total_injuries = 0

# Create a summary sheet
summary_sheet = wb.create_sheet(title="Summary")
summary_sheet.append(["Combined Injury Report Summary"])
summary_sheet["A1"].font = Font(bold=True, size=16)
summary_sheet.append([])
summary_sheet.append(["League", "Number of Teams with Injuries", "Total Injuries"])
summary_row = 4  # Start after the header

# Process each league
for league, file in league_files.items():
    if os.path.exists(file):
        try:
            # Read CSV data
            df = pd.read_csv(file)
            
            # Skip if empty
            if df.empty:
                print(f"⚠️ {league} file exists but contains no data")
                continue
                
            # Create a new sheet for the league
            ws = wb.create_sheet(title=league)
            
            # Add league name as a bold header
            ws.append([f"{league} Injury Report"])
            ws["A1"].font = Font(bold=True, size=16)
            ws.append([])  # Blank row
            
            # Get unique teams
            teams = sorted(df["Team"].unique())
            teams_with_injuries = len(teams)
            league_injuries = len(df)
            
            # Update total count
            total_injuries += league_injuries
            
            # Add to summary data
            available_leagues[league] = {
                "teams": teams_with_injuries,
                "injuries": league_injuries
            }
            
            row_num = 3  # Start from row 3 after league header
            for team in teams:
                team_data = df[df["Team"] == team]
                if team_data.empty:
                    continue
                    
                # Add team name as a title
                ws.append([team])
                ws.cell(row=row_num, column=1).font = Font(bold=True, size=14)
                row_num += 1
                
                # Add column headers
                headers = ["Player Name", "Injury Type", "Status", "Return Date", "Reported Date"]
                if "Short Comment" in team_data.columns and team_data["Short Comment"].any():
                    headers.append("Comment")
                
                ws.append(headers)
                for col_num, header in enumerate(headers, 1):
                    ws.cell(row=row_num, column=col_num).font = Font(bold=True)
                row_num += 1  # Move to the next row
                
                # Add team data
                for _, row in team_data.iterrows():
                    row_values = [
                        row["Player Name"], 
                        row["Injury Type"], 
                        row["Status"], 
                        row["Return Date"], 
                        row["Reported Date"]
                    ]
                    
                    # Add comment if it exists
                    if "Short Comment" in headers and "Short Comment" in row and row["Short Comment"]:
                        row_values.append(row["Short Comment"])
                    
                    ws.append(row_values)
                    row_num += 1
                
                # Add space before next team
                ws.append([])
                row_num += 1
                
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap width at 50 to avoid extremely wide columns
                ws.column_dimensions[col_letter].width = adjusted_width
                
            print(f"✅ Added {league} with {league_injuries} injuries across {teams_with_injuries} teams")
        
        except Exception as e:
            print(f"❌ Error processing {league}: {str(e)}")
    else:
        print(f"⚠️ {league} file not found: {file}")

# Fill in the summary sheet with the data we collected
for league, data in sorted(available_leagues.items()):
    summary_sheet.append([league, data["teams"], data["injuries"]])
    summary_row += 1

# Add total row
summary_sheet.append(["TOTAL", sum(data["teams"] for data in available_leagues.values()), total_injuries])
summary_sheet.cell(row=summary_row, column=1).font = Font(bold=True)
summary_sheet.cell(row=summary_row, column=2).font = Font(bold=True)
summary_sheet.cell(row=summary_row, column=3).font = Font(bold=True)

# Auto-adjust summary sheet column widths
for col in summary_sheet.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    summary_sheet.column_dimensions[col_letter].width = max_length + 2

# Save the Excel file
wb.save(output_file)
print(f"✅ Combined injury report saved as {output_file}")